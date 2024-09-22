import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import EPSC_App_Connection
import pandas as pd
import EPSC_preprocessing
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


# Streamlit UI
def main():
    if "multiple_pools_list" in st.session_state:
        st.session_state.multiple_pools_list = st.session_state.multiple_pools_list
    st.title("NSFA Analysis App")
    st.write("A simple webapp designed to run nonstationary fluctuation analysis on excitatory postsynaptic currents.")
    # File upload
    st.session_state.uploaded_files = st.file_uploader("Upload a file", type=["csv", "txt", "xlsx"], accept_multiple_files=True)
    if st.session_state.uploaded_files is not None:
        for idx, uploaded_file in enumerate(st.session_state.uploaded_files):
            if "filenameslist" in st.session_state:
                 if uploaded_file.name not in st.session_state["filenameslist"]:
                    st.session_state["filenameslist"].append(uploaded_file.name)
            else:
                st.session_state["filenameslist"] = []
                st.session_state["filenameslist"].append(uploaded_file.name)
            with st.expander(f"File: {uploaded_file.name}"):
                file_extension = uploaded_file.name.split('.')[-1]
                if file_extension == 'txt':
                    read_data = uploaded_file.read().decode("utf-8")
                    data = np.loadtxt(read_data.splitlines(), dtype=float)  # Specify dtype=float
                    st.session_state.EPSCs = data[:, :]

                if file_extension == 'csv':
                    st.session_state.EPSCs = pd.read_csv(uploaded_file)
                    st.session_state.EPSCs = st.session_state.EPSCs.to_numpy()

                if file_extension == 'xlsx':
                    sheet_number = st.number_input("Enter sheet number", min_value=0, step=1, key=str(idx) + "select")
                    st.session_state.EPSCs = pd.read_excel(uploaded_file, sheet_name=sheet_number)
                    st.session_state.EPSCs = st.session_state.EPSCs.to_numpy()

                else:
                    st.write("File-type not recognized")



                # Display the raw data
                st.subheader("Raw Data")
                st.write(st.session_state.EPSCs[0:10])

                # Parameter input
                time_duration = st.number_input("Enter time in ms", 5, key=str(idx) + "b")
                num_pools = st.slider("Enter number of pools", min_value=1, max_value=10, value=None, step=1,
                                      key=str(idx) + "c")
                # st.session_state.EPSCs = st.session_state.EPSCs_list[0]

                # Preprocessing
                st.subheader('Preprocessing')
                st.write("Preprocessing is optional.")
                st.session_state.use_processed = st.checkbox('Use Processed Data?',key = str(idx) + "useprocessed")

                # Set some essential variables
                num_samples = st.session_state.EPSCs.shape[0]
                timepoints, st.session_state.template = EPSC_App_Connection.create_template(st.session_state.EPSCs,
                                                                                            time_duration,
                                                                                            num_samples)
                peak_index = np.argmax(st.session_state.template)
                # Display data before processing
                st.write("Data before processing:")
                fig, axs = plt.subplots(1, 1)
                for i in range(st.session_state.EPSCs.shape[1]):
                    axs.plot(timepoints, st.session_state.EPSCs[:, i], label=f'Trace {i + 1}')

                axs.set_xlabel('Time (ms)', fontsize=13)
                axs.set_ylabel('Current (pA)', fontsize=13)
                axs.set_title('EPSCs Before Processing')
                st.pyplot(fig)
                st.write("For all parameters, leave as 0 to exclude criterion")
                invert = st.checkbox("Invert Trace", key=str(idx) + "checkInvert")
                peak_align = st.checkbox("Force peak alignment?", key=str(idx) + "checkPeak")
                max_rise_time = st.number_input("Enter max rise time in ms", 0, key=str(idx) + "d")
                min_peak_amp = st.number_input("Enter min peak amplitude in pA", 0, key=str(idx) + "e")
                pA_threshold = st.number_input("Enter tolerance of baseline return after peak in pA",
                                               key=str(idx) + "f")
                time_threshold = st.number_input("Enter how far along baseline to check for tolerance in ms",
                                                 key=str(idx) + "g")

                if invert:
                    st.session_state.EPSCs = st.session_state.EPSCs * -1
                if st.button("Preprocess Data", key=str(idx) + "preprocess"):
                    st.session_state.processed_data = None
                    if (max_rise_time != 0):
                        min_checked_data, count_peak = EPSC_preprocessing.check_minimum_peak_amplitude(
                            st.session_state.EPSCs, min_peak_amp, peak_index)
                    else:
                        min_checked_data = st.session_state.EPSCs
                        count_peak = 0
                    if peak_align:
                        aligned_data, count_aligned = EPSC_preprocessing.check_peak_alignment(min_checked_data,
                                                                                              peak_index)
                    else:
                        aligned_data = min_checked_data
                        count_aligned = 0
                    if (min_peak_amp != 0):
                        rise_data, count_rise = EPSC_preprocessing.check_rise_time(aligned_data, max_rise_time,
                                                                                   time_duration, peak_index)
                    else:
                        rise_data = aligned_data
                        count_rise = 0
                    if (pA_threshold != 0):
                        baseline_mean = EPSC_preprocessing.calculate_baseline_mean(rise_data, peak_index, time_duration)
                        data_returns_to_base, count_baseline = EPSC_preprocessing.check_return_to_base(rise_data,
                                                                                                       baseline_mean,
                                                                                                       time_duration,
                                                                                                       pA_threshold,
                                                                                                       time_threshold,
                                                                                                       peak_index)
                    else:
                        data_returns_to_base = rise_data
                        count_baseline = 0
                    st.session_state.processed_data = data_returns_to_base
                    fig, axs = plt.subplots(1, 1)
                    for i in range(st.session_state.processed_data.shape[1]):
                        axs.plot(timepoints, st.session_state.processed_data[:, i], label=f'Trace {i + 1}')

                    axs.set_xlabel('Time (ms)', fontsize=13)
                    axs.set_ylabel('Current (pA)', fontsize=13)
                    axs.set_title('EPSCs After Processing')
                    st.pyplot(fig)
                    df = pd.DataFrame({
                        'Condition': ['Minimum Peak Amplitude', 'Peak Alignment', 'Rise Time', 'Return to Baseline'],
                        'Traces Removed': [count_peak, count_aligned, count_rise, count_baseline]
                    })

                    st.write("### Traces Removed:")
                    st.table(df)

                # Analysis begins
                st.subheader('Analysis')

                # Button to trigger analysis
                if st.button("Create Template", key=str(idx) + "template"):
                    if st.session_state.use_processed:
                        st.session_state.EPSCs = st.session_state.processed_data
                        st.write("Processed data in use.")
                    else:
                        st.write("Unprocessed data in use.")
                    num_samples = st.session_state.EPSCs.shape[0]
                    timepoints, st.session_state.template = EPSC_App_Connection.create_template(st.session_state.EPSCs,
                                                                                                time_duration,
                                                                                                num_samples)
                    st.session_state.endPoint = st.session_state.template.shape[0] - 1
                    st.session_state.peak_index = np.argmax(st.session_state.template)

                    # Plot the result
                    st.subheader("Template Result")
                    st.write("Creating Template...")
                    st.write(f"Peak Value: {st.session_state.template[st.session_state.peak_index]}")
                    fig, axs = plt.subplots(2, 1)
                    axs[0].plot(timepoints, st.session_state.template)
                    axs[0].set_xlabel("Time(ms)")
                    axs[0].set_ylabel("Current (pA)")
                    axs[0].set_title("Average Template")
                    st.write("Processing traces by size... ")
                    trace_bins = EPSC_App_Connection.visualize_size_pools(st.session_state.EPSCs, num_pools)
                    print(st.session_state.EPSCs.shape)
                    for i in range(st.session_state.EPSCs.shape[1]):
                        axs[1].plot(timepoints, st.session_state.EPSCs[:, i], label=f'Trace {i + 1}',
                                    color=plt.cm.viridis(trace_bins[i] / num_pools))

                    axs[1].set_xlabel('Time (ms)', fontsize=13)
                    axs[1].set_ylabel('Current (pA)', fontsize=13)
                    axs[1].set_title('EPSCs Color-Coded by Peak Size Pools')
                    plt.tight_layout(h_pad=3)
                    st.pyplot(fig)

                if st.button("Run One Pool Analysis", key=str(idx) + "onepool"):
                    run_single_pool_analysis(num_pools)

                if st.button("Run Multiple Pool Analysis", key=str(idx) + "multipool"):
                    run_multipool_analysis(num_pools)


###SIDEBAR LOGIC

       # Sidebar for bulk report generation
        with st.sidebar:
            st.subheader("Bulk Report Generation")
            st.write("Upload files before selecting settings.")

            if "uploaded_files" not in st.session_state or not st.session_state.uploaded_files:
                st.write("You must upload at least one file to generate a report.")
            else:
                st.session_state.num_pools = st.number_input("Number of Pools",0)
                st.session_state.time_duration = st.number_input("Time Duration (ms)",0)

                # Display uploaded files and select sheets for XLSX files
                st.write("Select sheets for XLSX files:")

                # Dictionary to store selected sheets for each XLSX file
                if 'selected_sheets' not in st.session_state:
                    st.session_state.selected_sheets = {}

                for idx, uploaded_file in enumerate(st.session_state.uploaded_files):
                    file_name = uploaded_file.name
                    st.write(f"Processing: {file_name}")

                    # Handle XLSX files
                    if file_name.endswith(".xlsx"):
                        excel_file = pd.ExcelFile(uploaded_file)
                        sheet_names = excel_file.sheet_names

                        # Initialize selected sheets if not already in session state
                        if file_name not in st.session_state.selected_sheets:
                            st.session_state.selected_sheets[file_name] = {sheet: False for sheet in sheet_names}

                        for sheet in sheet_names:
                            selected = st.checkbox(f'{file_name}: Sheet {sheet}',
                                                   key=f'fileused_{file_name}_Sheet_{sheet}',
                                                   value=st.session_state.selected_sheets[file_name][sheet])
                            st.session_state.selected_sheets[file_name][sheet] = selected

                # Action button to run the processing
                if st.button("Run Processing"):
                    st.session_state.selected_files = []

                    for uploaded_file in st.session_state.uploaded_files:
                        file_name = uploaded_file.name
                        file_extension = file_name.split('.')[-1]

                        if file_extension == 'xlsx':
                            # Collect selected sheets for the XLSX file
                            selected_sheets = [sheet for sheet, is_selected in
                                               st.session_state.selected_sheets[file_name].items() if is_selected]
                            if selected_sheets:
                                st.session_state.selected_files.append((uploaded_file, 'xlsx', selected_sheets))

                        elif file_extension in ['csv', 'txt']:
                            # For non-XLSX files, use an empty list for consistency
                            st.session_state.selected_files.append((uploaded_file, file_extension, []))


                    # Process only the selected files and sheets
                    for uploaded_file, file_extension, selected_sheets in st.session_state.selected_files:
                        file_name = uploaded_file.name

                        if file_extension == 'xlsx':
                            st.write(f"Running action on XLSX file: {file_name} for sheets: {selected_sheets}")
                            excel_file = pd.ExcelFile(uploaded_file)

                            for sheet in selected_sheets:
                                df = pd.read_excel(excel_file, sheet_name=sheet)
                                st.write(f"Processing sheet: {sheet} in {file_name}")
                                st.session_state.EPSCs = df.to_numpy()
                                run_full_analysis(int(st.session_state.num_pools), int(st.session_state.time_duration))

                        elif file_extension == 'csv':
                            st.write(f"Running action on CSV file: {file_name}")
                            set_EPSCs(uploaded_file, "csv")
                            run_full_analysis(int(st.session_state.num_pools), int(st.session_state.time_duration))

                        elif file_extension == 'txt':
                            st.write(f"Running action on TXT file: {file_name}")
                            set_EPSCs(uploaded_file, "txt")
                            run_full_analysis(int(st.session_state.num_pools), int(st.session_state.time_duration))

                    st.success("Processing completed!")

        # Ensure multiplepools and singlepool_dict_list are in session state
        if "multiplepools" in st.session_state and "singlepool_dict_list" in st.session_state:
            st.write("Generating report...")
            flnme = "Report.xlsx"
            buffer = BytesIO()

            # Save data to the ExcelWriter and ensure buffer is correctly formatted
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Add headers to the "All Files Summary" sheet
                # One Pool header
                header = pd.DataFrame({"One Pool Information": ""}, index=[0])
                header.to_excel(writer, sheet_name='All Files Summary', startrow=0, startcol=0, index=False)
                single_categories = pd.DataFrame({"n": "", "i": ""}, index=[0])
                single_categories.to_excel(writer, sheet_name='All Files Summary', startrow=1, startcol=1, index=False)

                # Multi Pool header
                multi_header = pd.DataFrame({"Multi Pool Information": ""}, index=[0])
                multi_header.to_excel(writer, sheet_name='All Files Summary', startrow=0, startcol=6, index=False)
                multi_categories = pd.DataFrame({"Sheet Name": "", "Pool Number": "", "n": "", "i": ""}, index=[0])
                multi_categories.to_excel(writer, sheet_name='All Files Summary', startrow=1, startcol=5, index=False)

                # # Template header
                # template_header = pd.DataFrame({"Template Information": ""}, index=[0])
                # template_header.to_excel(writer, sheet_name='All Files Summary', startrow=0, startcol=12, index=False)
                # template_categories = pd.DataFrame({"Peak Value": ""}, index=[0])
                # template_categories.to_excel(writer, sheet_name='All Files Summary', startrow=1, startcol=12, index=False)

                # Loop through each single-pool item, now syncing with sheet names
                for idx, item in enumerate(st.session_state["singlepool_dict_list"]):
                    if idx < len(st.session_state.selected_files):
                        uploaded_file, file_extension, selected_sheets = st.session_state.selected_files[idx]

                        file_name = uploaded_file.name

                        if file_extension == 'xlsx':
                            for sheet in selected_sheets:
                                item = st.session_state["singlepool_dict_list"][idx]
                                # Write single pool information to individual sheets
                                header = pd.DataFrame({"One Pool Information": ""}, index=[0])
                                header.to_excel(writer, sheet_name=sheet, startrow=0, startcol=0)
                                df = pd.DataFrame(item, index=[0])
                                df.to_excel(writer, sheet_name=sheet, startrow=1, startcol=0)

                                # Add data to summary sheet
                                df.insert(0, 'Sheet Name', f"{file_name} - {sheet}")
                                df.to_excel(writer, sheet_name='All Files Summary', startrow=idx + 2, startcol=0,
                                            header=False, index=False)
                                idx+=1

                        elif file_extension == 'csv':
                            header = pd.DataFrame({"One Pool Information": ""}, index=[0])
                            header.to_excel(writer, sheet_name=file_name, startrow=0, startcol=0)
                            df = pd.DataFrame(item, index=[0])
                            df.to_excel(writer, sheet_name=file_name, startrow=1, startcol=0)
                            # Add data to summary sheet
                            df.insert(0, 'Sheet Name', f"{file_name}")
                            df.to_excel(writer, sheet_name='All Files Summary', startrow=idx + 2, startcol=0, header=False,
                                        index=False)

                        elif file_extension == 'txt':
                            header = pd.DataFrame({"One Pool Information": ""}, index=[0])
                            header.to_excel(writer, sheet_name=file_name, startrow=0, startcol=0)
                            df = pd.DataFrame(item, index=[0])
                            df.to_excel(writer, sheet_name=file_name, startrow=1, startcol=0)
                            # Add data to summary sheet
                            df.insert(0, 'Sheet Name', f"{file_name}")
                            df.to_excel(writer, sheet_name='All Files Summary', startrow=idx + 2, startcol=0, header=False,
                                        index=False)

                # Loop through each multi-pool item, now syncing with sheet names
                for idx, item in enumerate(st.session_state["multiplepools"]):
                    if idx < len(st.session_state.selected_files):
                        uploaded_file, file_extension, selected_sheets = st.session_state.selected_files[idx]
                        file_name = uploaded_file.name

                        if file_extension == 'xlsx':
                            for sheet in selected_sheets:
                                item = st.session_state["multiplepools"][idx]
                                # Write multi pool information to individual sheets
                                header = pd.DataFrame({"Multiple Pool Information": ""}, index=[0])
                                header.to_excel(writer, sheet_name=sheet, startrow=29, startcol=0)
                                item.to_excel(writer, sheet_name=sheet, startrow=30, startcol=0)

                                # Add data to summary sheet
                                # item.insert(0, 'Sheet Name', f"{file_name} - {sheet}")
                                item.to_excel(writer, sheet_name='All Files Summary',
                                              startrow=(idx) * st.session_state.num_pools + 2, startcol=6,
                                              header=False, index=False)
                                idx +=1

            # Reset buffer to the beginning before loading it with load_workbook
            buffer.seek(0)

            # Load the workbook correctly
            workbook = load_workbook(buffer)

            # Add images to the workbook, syncing with file names and sheet names
            for idx, (uploaded_file, file_extension, selected_sheets) in enumerate(st.session_state.selected_files):
                file_name = uploaded_file.name

                for sheet in selected_sheets:
                    # Handle single-pool plot
                    image_stream = BytesIO()
                    single_plot_fig = st.session_state["singlepool_plot_list"][idx]
                    single_plot_fig.savefig(image_stream, format='png')
                    image_stream.seek(0)
                    ws = workbook[sheet]
                    img = Image(image_stream)
                    img.height /= 1.2
                    img.width /= 1.2
                    ws.add_image(img, 'G1')

                    # Handle multi-pool plot
                    image_stream = BytesIO()
                    multi_pool_plot = st.session_state["multiplepools_plot_list"][idx]
                    multi_pool_plot.savefig(image_stream, format='png')
                    image_stream.seek(0)
                    img = Image(image_stream)
                    img.height /= 4
                    img.width /= 4
                    ws.add_image(img, 'G30')
                    idx +=1

                # Save the final Excel file
                final_excel_stream = BytesIO()
                workbook.save(final_excel_stream)
                final_excel_stream.seek(0)

                # Provide a download button
                st.download_button(label="Download Report", data=final_excel_stream, file_name=flnme,
                                   mime="application/vnd.ms-excel")



def close_factors(number):
    '''
    find the closest pair of factors for a given number
    '''
    factor1 = 0
    factor2 = number
    while factor1 + 1 <= factor2:
        factor1 += 1
        if number % factor1 == 0:
            factor2 = number // factor1

    return factor1, factor2


def to_excel(df, sheet_name):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name=sheet_name)
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    format1 = workbook.add_format({'num_format': '0.00'})
    worksheet.set_column('A:A', None, format1)
    writer.close()
    processed_data = output.getvalue()
    return processed_data


def run_full_analysis(num_pools,time_duration): #A catch-all function to run multipool, single, and template analysis
    num_samples = st.session_state.EPSCs.shape[0]
    timepoints, st.session_state.template = EPSC_App_Connection.create_template(st.session_state.EPSCs,
                                                                                time_duration,
                                                                                num_samples)
    st.session_state.endPoint = st.session_state.template.shape[0] - 1
    st.session_state.peak_index = np.argmax(st.session_state.template)
    run_multipool_analysis(num_pools=num_pools,display_plots=False,report_generation=True)
    run_single_pool_analysis(num_pools=num_pools, display_plots=False, report_generation=True)



def run_multipool_analysis(num_pools,display_plots=True,report_generation=False):
    if st.session_state.peak_index is None:
        st.write("Please create template first")
    else:
        pool_indices = EPSC_App_Connection.create_pool_indices(st.session_state.EPSCs,
                                                               st.session_state.peak_index, num_pools)
        raw_sorted = EPSC_App_Connection.sort_EPSCs_by_size(st.session_state.EPSCs,
                                                            st.session_state.peak_index)
        num_traces = st.session_state.EPSCs.shape[1]
        segment_indices, residuals_array = EPSC_App_Connection.create_segment_indices(num_traces,
                                                                                      raw_sorted,
                                                                                      st.session_state.template)
        multi_means_list, multi_var_list = EPSC_App_Connection.multi_pool_analysis(num_pools,
                                                                                   pool_indices,
                                                                                   raw_sorted,
                                                                                   residuals_array,
                                                                                   st.session_state.peak_index,
                                                                                   st.session_state.endPoint,
                                                                                   segment_indices)

        columns, rows = close_factors(num_pools)
        # Set the size of the figure
        fig, axs = plt.subplots(rows, columns,
                                figsize=(30, 20))  # Adjust the width and height as needed

        # Flatten the axs array if there is more than one row
        axs = axs.flatten()
        table_data = []

        # Create separate plots for each pair of means and variances
        for i in range(num_pools):
            # print("Pool length:", num_pools)
            # print("i", i)
            # print("var length:", len(multi_var_list))
            means = np.array(multi_means_list[i])
            variances = np.array(multi_var_list[i])
            coefficients = np.polyfit(means, variances, 2)
            coefficients[2] = 0
            if coefficients[0] > 0:
                coefficients = np.polyfit(means, variances, 1)
                coefficients[1] = 0
            fit_parab = np.poly1d(coefficients)
            roots = np.roots(coefficients)
            axs[i].scatter(means, variances, label=f'Pool {i + 1}')
            axs[i].set_title(f'Pool {i + 1}', fontsize=24, fontweight='bold')
            axs[i].set_xlabel('Mean Current (pA)', fontsize=20)
            axs[i].set_ylabel('Variances', fontsize=24)
            sorter = np.sort(means)
            axs[i].plot(sorter, fit_parab(sorter), color='black', label='Parabolic Fit')
            derivative_coefficients = np.polyder(coefficients)

            # Create a function representing the derivative
            derivative_function = np.poly1d(derivative_coefficients)

            # Calculate the initial slope at a specific point (e.g., x=1)
            initial_slope = derivative_function(0)
            # Calculate the n values
            n = roots[0] / initial_slope

            table_data.append([f'{i + 1}', n, initial_slope])
        df = pd.DataFrame(table_data, columns=['Pool Number', 'n', 'i'])
        if display_plots == True:
            st.pyplot(fig)
            st.table(df)

        if report_generation:
            if "multiplepools" not in st.session_state:
                st.session_state["multiplepools"] = []
                st.session_state["multiplepools_plot_list"] = []
                st.session_state["multiplepools_plot_list"].append(fig)

                st.session_state["multiplepools"].append(df)
            else:
                for item in st.session_state["multiplepools"]:
                    if df.equals(item):
                        use = False
                    else:
                        use = True
                if use == True:
                    st.session_state["multiplepools"].append(df)
                    st.session_state["multiplepools_plot_list"].append(fig)
def set_EPSCs(uploaded_file, file_extension):
    if file_extension == 'txt':
        uploaded_file.seek(0)  # Reset file pointer before reading
        read_data = uploaded_file.read().decode("utf-8")
        data = np.loadtxt(read_data.splitlines(), dtype=float)
        st.session_state.EPSCs = data[:, :]

    elif file_extension == 'csv':
        uploaded_file.seek(0)  # Reset file pointer before reading
        st.session_state.EPSCs = pd.read_csv(uploaded_file)
        st.session_state.EPSCs = st.session_state.EPSCs.to_numpy()

    elif file_extension == 'xlsx':
        sheet_number = st.number_input("Enter sheet number", min_value=0, step=1, key=str(uploaded_file.name) + "select")
        uploaded_file.seek(0)  # Reset file pointer before reading
        st.session_state.EPSCs = pd.read_excel(uploaded_file, sheet_name=sheet_number)
        st.session_state.EPSCs = st.session_state.EPSCs.to_numpy()

    # st.write("Finished Setting EPSCS")


def run_single_pool_analysis(num_pools,display_plots=True,report_generation=False):
    if st.session_state.peak_index is None:
        st.write("Please create template first")
    else:
        # st.write("Creating graph of one pool...")
        means, vars = EPSC_App_Connection.one_pool_analysis(st.session_state.EPSCs,
                                                            st.session_state.peak_index, num_pools,
                                                            st.session_state.endPoint,
                                                            st.session_state.template)

        # Fitting function
        fit_parabola, roots, initial_slope = EPSC_App_Connection.fitting_parabola(means, vars)

        n = roots[0] / initial_slope

        fig, axs = plt.subplots(1, 1)
        axs.scatter(means, vars, color='black')
        sorter = np.sort(means)
        axs.plot(sorter, fit_parabola(sorter), color='black')
        axs.set_title("Variance vs Mean")
        axs.set_xlabel("Mean Current (pA)")
        axs.set_ylabel("Current variance (pA^2)")
        if display_plots == True:
            st.pyplot(fig)
            st.write("Initial Current (i):", initial_slope)
            st.write("Number of Channels (n):", n)

        single_pool_dict = {"n": n, "i": initial_slope, "template_peak":st.session_state.template[st.session_state.peak_index]}
        if report_generation:
            if "singlepool_dict_list" not in st.session_state:
                st.session_state["singlepool_dict_list"] = []
                st.session_state["singlepool_plot_list"] = []
                st.session_state["singlepool_plot_list"].append(fig)
                st.session_state["singlepool_dict_list"].append(single_pool_dict)
            else:
                for item in st.session_state["singlepool_dict_list"]:
                    if single_pool_dict == item:
                        use = False
                    else:
                        use = True
                if use == True:
                    st.session_state["singlepool_dict_list"].append(single_pool_dict)
                    st.session_state["singlepool_plot_list"].append(fig)




if __name__ == "__main__":
    main()
